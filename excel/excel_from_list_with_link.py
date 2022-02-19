import openpyxl
import requests
from bs4 import BeautifulSoup


def write_excel_template_customize_link(filename, listdata):
    excel_file = openpyxl.Workbook()
    excel_sheet = excel_file.active

    # 타이틀 행은 여기에 적어주세요 예: excel_sheet.append(['', '제목', '링크'])
    # A열, B열등 각 열의 사이즈는 여기에 적어주세요 예: excel_sheet.column_dimensions['A'].width = 80
    excel_sheet.append([])
    excel_sheet.append(['', '닥터 아파트 분양 기사 타이틀'])
    excel_sheet.column_dimensions['B'].width = 80
    ## 선 정의
    thin = openpyxl.styles.Side(border_style="thin", color="01579B")

    for index, item in enumerate(listdata):
        excel_sheet.append(['', item[0]])
        excel_sheet.cell(row=index + 3, column=2).hyperlink = item[1]
        excel_sheet.cell(row=index + 3, column=2).border = openpyxl.styles.Border(top=thin, left=thin, right=thin,
                                                                                  bottom=thin)

    # 작성된 데이터를 기반으로 각 셀의 폰트 굵기/색, border, 정렬, 백그라운드 색은 여기에 적어주세요
    # 예:

    excel_sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center')  # 중앙정렬하기
    excel_sheet['B2'].font = openpyxl.styles.Font(size=14, b=True, color="01579B")  # 폰트 색깔 바꾸기
    excel_sheet['B2'].fill = openpyxl.styles.PatternFill(start_color="FFFDE7", fill_type="solid")
    excel_sheet['B2'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

    excel_file.save(filename)
    excel_file.close()


def crawling_with_link(url, css_selector, pre_url):
    return_data = list()
    res = requests.get(url)
    soup = BeautifulSoup(res.content, 'html.parser')
    datas1 = soup.select(css_selector)
    for item in datas1:
        return_data.append([item.get_text(), pre_url + item['href']])
    return return_data

datas1 = crawling_with_link('http://www.drapt.com/e_sale/index.htm?page_name=esale_news&menu_key=34', 'a.c0000000', 'http://www.drapt.com/e_sale/')
print(datas1)

write_excel_template_customize_link('tmp.xlsx', datas1)