import openpyxl


'''
excel_file = openpyxl.load_workbook('shop_in_seoul.xlsx')

# 쉬트 이름 확인 (리스트 타입으로 리턴됨)
excel_file.sheetnames

print("엑셀 sheet name:",excel_file.sheetnames)

# sheet 안에 있는 데이터 읽기
excel_sheet = excel_file["fastfood"]
#excel_sheet = excel_file.active #


for item in excel_sheet.rows:
    print(item[0].value, item[1].value, item[2].value)
'''

def read_excel_template(filename, sheetname):
    excel_file = openpyxl.load_workbook(filename)

    if sheetname == '':
        excel_sheet = excel_file.active
    else:
        excel_sheet = excel_file[sheetname]

    return_data = list()
    for item in excel_sheet.rows:
        datas1 = list()
        for item2 in item:
            datas1.append(item2.value)
        return_data.append(datas1)

    excel_file.close()
    return return_data


datas1 = read_excel_template('shop_in_seoul.xlsx', '')

for item in datas1:
  print(item)
